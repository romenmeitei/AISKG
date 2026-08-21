"""Deterministic replay of the AISKG v3.1.2 additional analyses.

This module deliberately separates two reproducibility levels:

1. deterministic statistical replay from archived item-level pathway and
   benchmark outputs (the default and manuscript-facing route); and
2. the historical external PubTator/LLM execution, preserved as an executed
   reference notebook but not presented as bit-for-bit rerunnable because the
   LLM run recorded the mutable revision ``main`` rather than a commit SHA.
"""
from __future__ import annotations

import json
import math
import re
import shutil
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.optimize import linear_sum_assignment
from scipy.stats import binomtest, fisher_exact, norm
from statsmodels.stats.multitest import multipletests

from ..utils import FIXED_ZIP_TIMESTAMP, deterministic_zip, sha256_file, write_sha256s
from .reviewer_validation import replay_reviewer_validation

RELEASE_VERSION = "3.1.2"
BASE_FROZEN_RELEASE = "3.0.0"
AUDITED_COMMIT = "0e9e0e979c98664c74d7f27e318a7a06aed4fa54"
DEFAULT_SEED = 20260817
DEFAULT_PATHWAY_BOOTSTRAPS = 10_000
DEFAULT_BENCHMARK_BOOTSTRAPS = 5_000

RATING_COLUMNS = [
    "entities_correct",
    "relations_correct",
    "direction_correct",
    "sequence_coherent",
    "terminal_class_correct",
    "all_edges_evidence_supported",
    "complete_pathway_correct",
]
PRIMARY_ENDPOINT = "complete_pathway_correct"
COMMON_ENTITY_TYPES = {"CHEMICAL", "DISEASE", "GENE_PROTEIN", "SPECIES", "CELL_LINE", "VARIANT"}


@dataclass(frozen=True)
class ReplayResult:
    """Paths and key values returned by :func:`run_replay`."""

    output_root: Path
    output_archive: Path
    pathway_summary: Path
    benchmark_metrics: Path
    success_marker: Path
    manifest: Path


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _verify_input_checksums(data_root: Path) -> int:
    checksum_file = data_root / "SHA256SUMS.txt"
    if not checksum_file.exists():
        raise FileNotFoundError(f"Missing frozen-input checksum manifest: {checksum_file}")
    checked = 0
    for line in checksum_file.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        expected, relative = line.split("  ", 1)
        path = data_root / relative
        if not path.is_file():
            raise FileNotFoundError(f"Missing frozen additional-analysis input: {relative}")
        observed = sha256_file(path)
        if observed != expected:
            raise ValueError(f"Frozen additional-analysis checksum mismatch for {relative}: {observed} != {expected}")
        checked += 1
    return checked


def _write_json(payload: Mapping[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True, default=str) + "\n", encoding="utf-8")


def _normalise_xlsx_archive(path: Path) -> None:
    """Rewrite an XLSX container with fixed metadata and ZIP timestamps.

    OpenPyXL writes current timestamps into the OOXML container by default.
    Normalising both the workbook core properties and every ZIP member makes
    the generated workbook byte-for-byte reproducible across repeated runs.
    """

    temporary = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(path, "r") as source:
        members = [(info.filename, source.read(info.filename), info.compress_type) for info in source.infolist()]
    with zipfile.ZipFile(temporary, "w") as target:
        for filename, payload, compression in sorted(members, key=lambda item: item[0]):
            if filename == "docProps/core.xml":
                core = payload.decode("utf-8")
                for tag in ("created", "modified"):
                    core = re.sub(
                        rf"(<dcterms:{tag}\b[^>]*>).*?(</dcterms:{tag}>)",
                        rf"\g<1>2026-01-01T00:00:00Z\g<2>",
                        core,
                    )
                payload = core.encode("utf-8")
            info = zipfile.ZipInfo(filename, date_time=FIXED_ZIP_TIMESTAMP)
            info.compress_type = compression
            info.create_system = 3
            info.external_attr = 0o644 << 16
            target.writestr(
                info,
                payload,
                compress_type=compression,
                compresslevel=6 if compression == zipfile.ZIP_DEFLATED else None,
            )
    temporary.replace(path)


def _set_fixed_workbook_properties(writer: pd.ExcelWriter) -> None:
    fixed = datetime(*FIXED_ZIP_TIMESTAMP)
    writer.book.properties.created = fixed
    writer.book.properties.modified = fixed
    writer.book.properties.creator = "AISKG reproducibility pipeline"
    writer.book.properties.lastModifiedBy = "AISKG reproducibility pipeline"


def _style_excel_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    frame: pd.DataFrame,
    *,
    width_overrides: Mapping[str, float] | None = None,
    wrap_columns: Iterable[str] = (),
    max_width: float = 34.0,
) -> None:
    """Apply deterministic, reviewer-friendly formatting to an exported table.

    Values remain unrounded in the workbook; numeric formats only control their
    displayed precision. Widths are bounded so long evidence fields wrap rather
    than making the worksheet impractically wide.
    """

    worksheet = writer.sheets[sheet_name]
    overrides = dict(width_overrides or {})
    wrapped = set(wrap_columns)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 30

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for column_index, column_name in enumerate(frame.columns, start=1):
        column_letter = get_column_letter(column_index)
        series = frame[column_name]
        if column_name in overrides:
            width = float(overrides[column_name])
        elif pd.api.types.is_numeric_dtype(series):
            width = max(12.0, min(max_width, len(str(column_name)) + 3.0))
        else:
            lengths = [len(str(column_name))]
            lengths.extend(len(str(value)) for value in series.dropna().astype(str).head(250))
            width = max(12.0, min(max_width, max(lengths, default=10) + 2.0))
        worksheet.column_dimensions[column_letter].width = width

        is_float = pd.api.types.is_float_dtype(series)
        is_integer = pd.api.types.is_integer_dtype(series)
        should_wrap = column_name in wrapped or width >= 30
        for cell in worksheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=2,
            max_row=max(2, len(frame) + 1),
        ):
            for item in cell:
                item.font = body_font
                item.alignment = Alignment(
                    horizontal="left" if should_wrap or not (is_float or is_integer) else "right",
                    vertical="top",
                    wrap_text=should_wrap,
                )
                if is_float:
                    item.number_format = "0.000"
                elif is_integer:
                    item.number_format = "0"


def _style_excel_workbook(
    writer: pd.ExcelWriter,
    frames: Mapping[str, pd.DataFrame],
    *,
    sheet_options: Mapping[str, Mapping[str, Any]] | None = None,
) -> None:
    options = dict(sheet_options or {})
    for sheet_name, frame in frames.items():
        _style_excel_sheet(writer, sheet_name, frame, **dict(options.get(sheet_name, {})))


def _prf_from_counts(counts: pd.DataFrame) -> dict[str, float | int]:
    tp, fp, fn = counts[["tp", "fp", "fn"]].sum()
    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    f1 = 2 * precision * recall / (precision + recall) if precision + recall else 0.0
    return {
        "tp": int(tp),
        "fp": int(fp),
        "fn": int(fn),
        "precision": float(precision),
        "recall": float(recall),
        "f1": float(f1),
    }


def _wilson_interval(k: int, n: int, alpha: float = 0.05) -> tuple[float, float]:
    if n == 0:
        return math.nan, math.nan
    z = float(norm.ppf(1 - alpha / 2))
    proportion = k / n
    denominator = 1 + z * z / n
    center = (proportion + z * z / (2 * n)) / denominator
    half = z * math.sqrt(proportion * (1 - proportion) / n + z * z / (4 * n * n)) / denominator
    return float(center - half), float(center + half)


def _pathway_arm(frame: pd.DataFrame, arm: str) -> pd.DataFrame:
    if arm == "PRE_REFINEMENT":
        return frame[frame["membership"].isin(["SHARED", "REMOVED_BY_REFINEMENT"])].copy()
    if arm == "OUTCOME_AWARE_REFINED":
        return frame[frame["membership"].isin(["SHARED", "ADDED_BY_REFINEMENT"])].copy()
    raise ValueError(f"Unknown pathway arm: {arm}")


def _normalise_comparison_value(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, (np.integer, int)):
        return str(int(value))
    if isinstance(value, (np.floating, float)) and float(value).is_integer():
        return str(int(value))
    return str(value)


def _assert_frame_close(
    observed: pd.DataFrame,
    expected: pd.DataFrame,
    key_columns: Sequence[str],
    *,
    tolerance: float = 1e-12,
) -> None:
    """Verify tables while tolerating CSV dtype changes in mixed key columns."""

    if list(observed.columns) != list(expected.columns):
        raise AssertionError(f"Column mismatch: {list(observed.columns)} != {list(expected.columns)}")
    if len(observed) != len(expected):
        raise AssertionError(f"Row-count mismatch: {len(observed)} != {len(expected)}")

    def indexed(frame: pd.DataFrame) -> pd.DataFrame:
        working = frame.copy()
        helper_columns: list[str] = []
        for column in key_columns:
            helper = f"__key_{column}"
            working[helper] = working[column].map(_normalise_comparison_value)
            helper_columns.append(helper)
        return working.sort_values(helper_columns, kind="stable").drop(columns=helper_columns).reset_index(drop=True)

    observed_sorted = indexed(observed)
    expected_sorted = indexed(expected)
    for column in observed.columns:
        if pd.api.types.is_numeric_dtype(observed_sorted[column]) and pd.api.types.is_numeric_dtype(expected_sorted[column]):
            left = observed_sorted[column].to_numpy(dtype=float)
            right = expected_sorted[column].to_numpy(dtype=float)
            if not np.allclose(left, right, rtol=0.0, atol=tolerance, equal_nan=True):
                max_delta = float(np.nanmax(np.abs(left - right)))
                raise AssertionError(f"Numeric mismatch in {column}; max absolute delta={max_delta}")
        else:
            left = observed_sorted[column].map(_normalise_comparison_value)
            right = expected_sorted[column].map(_normalise_comparison_value)
            if not left.equals(right):
                raise AssertionError(f"Text mismatch in {column}")


def _validate_pathway_input(pathway: pd.DataFrame) -> None:
    required = {
        "validation_id",
        "membership",
        "path_text",
        "pathway_template",
        "n_edges",
        "contains_intervention",
        "terminal_class",
        *RATING_COLUMNS,
        *[f"{column}_source" for column in RATING_COLUMNS],
        *[f"{column}_binary" for column in RATING_COLUMNS],
    }
    missing = required - set(pathway.columns)
    if missing:
        raise ValueError(f"Missing pathway columns: {sorted(missing)}")
    if len(pathway) != 115 or not pathway["validation_id"].is_unique:
        raise ValueError("Expected exactly 115 unique pathway review units.")
    membership = pathway["membership"].value_counts().to_dict()
    expected_membership = {"REMOVED_BY_REFINEMENT": 63, "SHARED": 32, "ADDED_BY_REFINEMENT": 20}
    if membership != expected_membership:
        raise ValueError(f"Unexpected pathway membership counts: {membership}")
    for column in RATING_COLUMNS:
        labels = set(pathway[column].dropna().astype(str).unique())
        if not labels <= {"Yes", "No"}:
            raise ValueError(f"Non-binary final labels in {column}: {sorted(labels)}")
        binary = pathway[f"{column}_binary"]
        if not binary.isin([0, 1]).all():
            raise ValueError(f"Invalid binary values in {column}_binary")
        expected = pathway[column].eq("Yes").astype(int)
        if not binary.astype(int).equals(expected):
            raise ValueError(f"Label/binary inconsistency in {column}")


def _replay_pathway(
    data_dir: Path,
    output_dir: Path,
    *,
    seed: int,
    bootstrap_iterations: int,
) -> dict[str, Any]:
    frozen_final = pd.read_csv(data_dir / "pathway_validation_final_labels_public.csv", dtype={"validation_id": str})
    protocol = _read_json(data_dir / "pathway_validation_public_protocol.json")
    workbook_dir = data_dir / "reviewer_workbooks"
    reviewer = replay_reviewer_validation(
        workbook_dir / "Expert_A_completed_public.xlsx",
        workbook_dir / "Expert_B_completed_public.xlsx",
        workbook_dir / "Third_Expert_completed_public.xlsx",
    )

    expected_agreement = pd.read_csv(data_dir / "pathway_interrater_agreement_expected.csv")
    _assert_frame_close(reviewer.agreement, expected_agreement, ["dimension"])
    expected_qc = _read_json(data_dir / "pathway_expected_reviewer_qc.json")
    observed_qc = {key: reviewer.qc[key] for key in expected_qc}
    if observed_qc != expected_qc:
        raise AssertionError("Recomputed reviewer-workbook QC does not match the frozen expected record.")

    reconstructed = reviewer.reconstructed_ratings.set_index("validation_id")
    pathway = frozen_final.copy()
    for column in reconstructed.columns:
        pathway[column] = pathway["validation_id"].map(reconstructed[column])
    _assert_frame_close(pathway, frozen_final, ["validation_id"])
    _validate_pathway_input(pathway)

    summary_rows: list[dict[str, Any]] = []
    for arm in ["PRE_REFINEMENT", "OUTCOME_AWARE_REFINED"]:
        subset = _pathway_arm(pathway, arm)
        for endpoint in RATING_COLUMNS:
            correct = int(subset[f"{endpoint}_binary"].sum())
            n = len(subset)
            low, high = _wilson_interval(correct, n)
            summary_rows.append(
                {
                    "system": arm,
                    "endpoint": endpoint,
                    "correct": correct,
                    "n": n,
                    "proportion": correct / n,
                    "CI_low": low,
                    "CI_high": high,
                }
            )
    summary = pd.DataFrame(summary_rows)

    # Vectorized membership masks preserve the original seeded bootstrap while
    # avoiding repeated DataFrame construction inside 10,000 replicates.
    rng = np.random.default_rng(seed + 1)
    membership_values = pathway["membership"].to_numpy(dtype=object)
    endpoint_values = pathway[f"{PRIMARY_ENDPOINT}_binary"].to_numpy(dtype=float)
    pre_membership = np.isin(membership_values, ["SHARED", "REMOVED_BY_REFINEMENT"])
    refined_membership = np.isin(membership_values, ["SHARED", "ADDED_BY_REFINEMENT"])
    delta_array = np.empty(bootstrap_iterations, dtype=float)
    for replicate in range(bootstrap_iterations):
        sampled_indices = rng.integers(0, len(pathway), len(pathway))
        sampled_values = endpoint_values[sampled_indices]
        pre_mask = pre_membership[sampled_indices]
        refined_mask = refined_membership[sampled_indices]
        delta_array[replicate] = sampled_values[refined_mask].mean() - sampled_values[pre_mask].mean()

    pre_estimate = float(endpoint_values[pre_membership].mean())
    refined_estimate = float(endpoint_values[refined_membership].mean())
    contrast = pd.DataFrame(
        [
            {
                "endpoint": PRIMARY_ENDPOINT,
                "pre_refinement": pre_estimate,
                "outcome_aware_refined": refined_estimate,
                "absolute_difference": refined_estimate - pre_estimate,
                "bootstrap_CI_low": float(np.quantile(delta_array, 0.025)),
                "bootstrap_CI_high": float(np.quantile(delta_array, 0.975)),
                "bootstrap_probability_improvement": float((delta_array > 0).mean()),
                "method": "Unique-path cluster bootstrap preserving overlapping system membership",
            }
        ]
    )

    mechanism_rows: list[dict[str, Any]] = []
    for membership, subset in pathway.groupby("membership"):
        correct = int(subset[f"{PRIMARY_ENDPOINT}_binary"].sum())
        n = len(subset)
        low, high = _wilson_interval(correct, n)
        mechanism_rows.append(
            {
                "membership": membership,
                "correct": correct,
                "n": n,
                "proportion": correct / n,
                "CI_low": low,
                "CI_high": high,
            }
        )
    mechanism = pd.DataFrame(mechanism_rows)

    removed = pathway.loc[pathway["membership"].eq("REMOVED_BY_REFINEMENT"), f"{PRIMARY_ENDPOINT}_binary"]
    retained = pathway.loc[pathway["membership"].eq("SHARED"), f"{PRIMARY_ENDPOINT}_binary"]
    selection_table = [
        [int(retained.sum()), int((1 - retained).sum())],
        [int(removed.sum()), int((1 - removed).sum())],
    ]
    odds_ratio, fisher_p = fisher_exact(selection_table)
    selection_test = pd.DataFrame(
        [
            {
                "comparison": "Shared/retained vs removed pathways",
                "odds_ratio_correctness": float(odds_ratio),
                "fisher_exact_P": float(fisher_p),
                "table_correct_incorrect": str(selection_table),
            }
        ]
    )

    expanded: list[pd.DataFrame] = []
    for arm in ["PRE_REFINEMENT", "OUTCOME_AWARE_REFINED"]:
        subset = _pathway_arm(pathway, arm)
        subset["system"] = arm
        expanded.append(subset)
    combined = pd.concat(expanded, ignore_index=True)
    stratified_rows: list[dict[str, Any]] = []
    for stratifier in ["pathway_template", "n_edges", "contains_intervention", "terminal_class"]:
        for (system, level), subset in combined.groupby(["system", stratifier], dropna=False):
            correct = int(subset[f"{PRIMARY_ENDPOINT}_binary"].sum())
            n = len(subset)
            low, high = _wilson_interval(correct, n)
            stratified_rows.append(
                {
                    "stratifier": stratifier,
                    "level": level,
                    "system": system,
                    "correct": correct,
                    "n": n,
                    "proportion": correct / n,
                    "CI_low": low,
                    "CI_high": high,
                }
            )
    stratified = pd.DataFrame(stratified_rows)

    _assert_frame_close(summary, pd.read_csv(data_dir / "pathway_expected_summary.csv"), ["system", "endpoint"])
    _assert_frame_close(contrast, pd.read_csv(data_dir / "pathway_expected_primary_contrast.csv"), ["endpoint"])
    _assert_frame_close(mechanism, pd.read_csv(data_dir / "pathway_expected_membership_correctness.csv"), ["membership"])
    _assert_frame_close(selection_test, pd.read_csv(data_dir / "pathway_expected_selection_test.csv"), ["comparison"])
    _assert_frame_close(
        stratified,
        pd.read_csv(data_dir / "pathway_expected_stratified.csv"),
        ["stratifier", "system", "level"],
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    pathway.to_csv(output_dir / "pathway_validation_final_labels_public.csv", index=False)
    pathway.to_csv(output_dir / "pathway_validation_final_labels_reconstructed.csv", index=False)
    reviewer.expert_a.to_csv(output_dir / "expert_A_completed_public.csv", index=False)
    reviewer.expert_b.to_csv(output_dir / "expert_B_completed_public.csv", index=False)
    reviewer.third_expert.to_csv(output_dir / "third_expert_completed_public.csv", index=False)
    reviewer.agreement.to_csv(output_dir / "interrater_agreement_recomputed.csv", index=False)
    reviewer.rating_matrix_long.to_csv(output_dir / "reviewer_rating_matrix_long.csv", index=False)
    reviewer.adjudication_audit.to_csv(output_dir / "third_expert_adjudication_audit.csv", index=False)
    _write_json(reviewer.qc, output_dir / "REVIEWER_WORKBOOK_QC.json")
    for filename in [
        "Expert_A_completed_public.xlsx",
        "Expert_B_completed_public.xlsx",
        "Third_Expert_completed_public.xlsx",
    ]:
        shutil.copy2(workbook_dir / filename, output_dir / filename)
    summary.to_csv(output_dir / "pathway_correctness_summary.csv", index=False)
    contrast.to_csv(output_dir / "primary_refinement_contrast.csv", index=False)
    mechanism.to_csv(output_dir / "refinement_membership_correctness.csv", index=False)
    selection_test.to_csv(output_dir / "refinement_selection_fisher_test.csv", index=False)
    stratified.to_csv(output_dir / "pathway_correctness_stratified.csv", index=False)

    primary = summary[summary["endpoint"].eq(PRIMARY_ENDPOINT)].copy()
    figure, axis = plt.subplots(figsize=(7, 5))
    x_positions = np.arange(len(primary))
    proportions = primary["proportion"].to_numpy()
    errors = np.vstack(
        [
            proportions - primary["CI_low"].to_numpy(),
            primary["CI_high"].to_numpy() - proportions,
        ]
    )
    axis.bar(x_positions, proportions, width=0.65)
    axis.errorbar(x_positions, proportions, yerr=errors, fmt="none", capsize=5)
    axis.set_xticks(x_positions, ["Pre-refinement", "Outcome-aware\nrefined"])
    axis.set_ylabel("Complete-pathway correctness")
    axis.set_ylim(0, 1)
    for index, row in primary.reset_index(drop=True).iterrows():
        axis.text(index, row.proportion + 0.04, f"{int(row.correct)}/{int(row.n)}\n({row.proportion:.1%})", ha="center")
    figure.tight_layout()
    figure.savefig(output_dir / "complete_pathway_correctness_comparison.png", dpi=300, bbox_inches="tight")
    plt.close(figure)

    membership_order = ["REMOVED_BY_REFINEMENT", "SHARED", "ADDED_BY_REFINEMENT"]
    ordered_mechanism = mechanism.set_index("membership").reindex(membership_order).reset_index()
    figure, axis = plt.subplots(figsize=(8, 5))
    axis.bar(range(len(ordered_mechanism)), ordered_mechanism["proportion"])
    axis.set_xticks(range(len(ordered_mechanism)), ["Removed", "Shared/retained", "Added"])
    axis.set_ylabel("Complete-pathway correctness")
    axis.set_ylim(0, 1)
    for index, row in ordered_mechanism.iterrows():
        axis.text(index, row.proportion + 0.035, f"{int(row.correct)}/{int(row.n)}", ha="center")
    figure.tight_layout()
    figure.savefig(output_dir / "refinement_mechanism_correctness.png", dpi=300, bbox_inches="tight")
    plt.close(figure)

    pre_row = primary[primary["system"].eq("PRE_REFINEMENT")].iloc[0]
    refined_row = primary[primary["system"].eq("OUTCOME_AWARE_REFINED")].iloc[0]
    contrast_row = contrast.iloc[0]
    complete_agreement = reviewer.agreement.loc[
        reviewer.agreement["dimension"].eq(PRIMARY_ENDPOINT)
    ].iloc[0]
    manuscript_text = (
        f"In a blinded census-based validation, {int(pre_row.correct)}/{int(pre_row.n)} "
        f"pre-refinement pathways ({pre_row.proportion:.1%}, Wilson 95% CI "
        f"{pre_row.CI_low:.1%}–{pre_row.CI_high:.1%}) and "
        f"{int(refined_row.correct)}/{int(refined_row.n)} outcome-aware refined pathways "
        f"({refined_row.proportion:.1%}, Wilson 95% CI {refined_row.CI_low:.1%}–"
        f"{refined_row.CI_high:.1%}) were completely correct. The absolute difference was "
        f"{contrast_row.absolute_difference:.1%} (overlap-aware cluster-bootstrap 95% CI "
        f"{contrast_row.bootstrap_CI_low:.1%}–{contrast_row.bootstrap_CI_high:.1%}). "
        "The evaluation covered all 95 historical and all 52 refined eligible pathway-plus-template "
        "items, represented by 115 unique blinded review items. Reviewer-level replay verified "
        f"805 paired ratings, {reviewer.qc['direct_A_B_disagreements']} direct A–B disagreements, "
        f"and all {reviewer.qc['required_third_expert_adjudications']} required third-expert decisions; "
        f"complete-pathway raw agreement was {complete_agreement.raw_agreement:.1%}, Cohen's kappa "
        f"{complete_agreement.cohen_kappa:.3f}, and Gwet's AC1 {complete_agreement.gwet_ac1:.3f}."
    )
    (output_dir / "manuscript_ready_results_with_provenance_note.txt").write_text(manuscript_text + "\n", encoding="utf-8")

    pathway_workbook = output_dir / "AISKG_Expanded_Pathway_Validation_Results_v3.1.2.xlsx"
    reviewer_qc_frame = pd.DataFrame(
        [
            (key, json.dumps(value, sort_keys=True) if isinstance(value, (dict, list)) else value)
            for key, value in reviewer.qc.items()
        ],
        columns=["QC_item", "value"],
    )
    pathway_sheets = {
        "Agreement_recomputed": reviewer.agreement,
        "Reviewer_QC": reviewer_qc_frame,
        "Third_Adjudication": reviewer.adjudication_audit,
        "Reviewer_Rating_Matrix": reviewer.rating_matrix_long,
        "System_Summary": summary,
        "Primary_Contrast": contrast,
        "Refinement_Mechanism": mechanism,
        "Selection_Test": selection_test,
        "Stratified": stratified,
        "Final_Public_Labels": pathway,
    }
    pathway_sheet_options = {
        "Agreement_recomputed": {
            "width_overrides": {"dimension": 34, "n": 10, "disagreements": 16},
            "max_width": 22,
        },
        "Reviewer_QC": {
            "width_overrides": {"QC_item": 45, "value": 85},
            "wrap_columns": ["QC_item", "value"],
            "max_width": 85,
        },
        "Third_Adjudication": {
            "width_overrides": {
                "validation_id": 14,
                "path_text": 42,
                "pathway_template": 34,
                "rating_dimension": 32,
                "expert_A_label": 16,
                "expert_B_label": 16,
                "expert_A_comment": 55,
                "expert_B_comment": 55,
                "adjudication_reason": 34,
                "adjudicated_label": 18,
                "adjudicator_rationale": 60,
                "final_binary": 13,
                "source_match_verified": 22,
            },
            "wrap_columns": [
                "path_text",
                "pathway_template",
                "rating_dimension",
                "expert_A_comment",
                "expert_B_comment",
                "adjudication_reason",
                "adjudicator_rationale",
            ],
            "max_width": 60,
        },
        "Reviewer_Rating_Matrix": {
            "width_overrides": {
                "validation_id": 14,
                "path_text": 42,
                "pathway_template": 34,
                "rating_dimension": 32,
                "expert_A_comment": 50,
                "expert_B_comment": 50,
                "adjudicator_rationale": 55,
            },
            "wrap_columns": [
                "path_text",
                "pathway_template",
                "rating_dimension",
                "expert_A_comment",
                "expert_B_comment",
                "adjudicator_rationale",
            ],
            "max_width": 55,
        },
        "System_Summary": {"max_width": 30},
        "Primary_Contrast": {"max_width": 26},
        "Refinement_Mechanism": {"max_width": 30},
        "Selection_Test": {"max_width": 28},
        "Stratified": {"max_width": 34},
        "Final_Public_Labels": {
            "width_overrides": {
                "validation_id": 14,
                "path_text": 42,
                "pathway_template": 34,
                "representative_evidence_sentences": 65,
            },
            "wrap_columns": [
                "path_text",
                "pathway_template",
                "representative_evidence_sentences",
            ],
            "max_width": 42,
        },
    }
    with pd.ExcelWriter(pathway_workbook, engine="openpyxl") as writer:
        _set_fixed_workbook_properties(writer)
        for sheet_name, frame in pathway_sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
        _style_excel_workbook(writer, pathway_sheets, sheet_options=pathway_sheet_options)
    _normalise_xlsx_archive(pathway_workbook)

    status = {
        "release": RELEASE_VERSION,
        "status": "REVIEWER_LEVEL_AND_FINAL_LABEL_REPLAY_PASSED",
        "unique_items": len(pathway),
        "paired_reviewer_ratings": int(reviewer.qc["total_A_B_rating_pairs"]),
        "direct_A_B_disagreements": int(reviewer.qc["direct_A_B_disagreements"]),
        "required_third_expert_adjudications": int(reviewer.qc["required_third_expert_adjudications"]),
        "third_expert_exact_coverage_verified": True,
        "final_labels_match_prior_public_table": True,
        "pre_refinement_correct": "23/95",
        "outcome_aware_refined_correct": "26/52",
        "all_expected_tables_verified": True,
        "completed_public_reviewer_workbooks_available": bool(
            protocol.get("public_expert_A_completed_workbook_available")
            and protocol.get("public_expert_B_completed_workbook_available")
            and protocol.get("public_third_expert_completed_workbook_available")
        ),
        "agreement_status": "INDEPENDENTLY_RECOMPUTED_FROM_PUBLIC_SANITIZED_WORKBOOKS",
        "source_rollup_exceptions_disclosed": reviewer.qc["complete_pathway_rollup_inconsistencies"],
    }
    _write_json(status, output_dir / "PATHWAY_REPRODUCIBILITY_STATUS.json")
    return {
        "summary": summary,
        "contrast": contrast,
        "mechanism": mechanism,
        "selection_test": selection_test,
        "stratified": stratified,
        "agreement": reviewer.agreement,
        "reviewer_qc": reviewer.qc,
        "adjudication_audit": reviewer.adjudication_audit,
        "status": status,
        "manuscript_text": manuscript_text,
    }

def _heldout_normalise_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).replace("α", "alpha").replace("β", "beta").replace("γ", "gamma")
    return re.sub(r"\s+", " ", text.casefold()).strip()


def _normalise_label(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"[^A-Z0-9]+", "_", str(value).upper()).strip("_")


def _max_entity_match(gold: pd.DataFrame, predictions: pd.DataFrame, mode: str = "strict") -> int:
    if len(gold) == 0 or len(predictions) == 0:
        return 0
    matrix = np.zeros((len(gold), len(predictions)), dtype=int)
    for gold_index, (_, gold_row) in enumerate(gold.reset_index(drop=True).iterrows()):
        for prediction_index, (_, prediction_row) in enumerate(predictions.reset_index(drop=True).iterrows()):
            if gold_row.type != prediction_row.type or pd.isna(gold_row.start) or pd.isna(prediction_row.start):
                continue
            if mode == "strict":
                match = int(gold_row.start) == int(prediction_row.start) and int(gold_row.end) == int(prediction_row.end)
            elif mode == "overlap":
                match = max(int(gold_row.start), int(prediction_row.start)) < min(int(gold_row.end), int(prediction_row.end))
            else:
                raise ValueError(f"Unsupported entity matching mode: {mode}")
            matrix[gold_index, prediction_index] = int(match)
    row_indices, column_indices = linear_sum_assignment(-matrix)
    return int(matrix[row_indices, column_indices].sum())


def _entity_counts(
    gold: pd.DataFrame,
    predictions: pd.DataFrame,
    classes: Iterable[str],
    sentence_ids: Sequence[str],
    mode: str,
) -> pd.DataFrame:
    classes = set(classes)
    rows: list[dict[str, Any]] = []
    for sentence_id in sentence_ids:
        gold_subset = gold[(gold["benchmark_sentence_id"].eq(sentence_id)) & gold["type"].isin(classes)]
        prediction_subset = predictions[
            (predictions["benchmark_sentence_id"].eq(sentence_id)) & predictions["type"].isin(classes)
        ]
        true_positives = _max_entity_match(gold_subset, prediction_subset, mode)
        rows.append(
            {
                "benchmark_sentence_id": sentence_id,
                "tp": true_positives,
                "fp": len(prediction_subset) - true_positives,
                "fn": len(gold_subset) - true_positives,
            }
        )
    return pd.DataFrame(rows)


def _relation_key(row: pd.Series) -> tuple[str, str, str]:
    return (
        _heldout_normalise_text(row.source),
        _normalise_label(row.relation),
        _heldout_normalise_text(row.target),
    )


def _relation_counts(gold: pd.DataFrame, predictions: pd.DataFrame, sentence_ids: Sequence[str]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for sentence_id in sentence_ids:
        gold_keys = {_relation_key(row) for _, row in gold[gold["benchmark_sentence_id"].eq(sentence_id)].iterrows()}
        prediction_keys = {
            _relation_key(row) for _, row in predictions[predictions["benchmark_sentence_id"].eq(sentence_id)].iterrows()
        }
        rows.append(
            {
                "benchmark_sentence_id": sentence_id,
                "tp": len(gold_keys & prediction_keys),
                "fp": len(prediction_keys - gold_keys),
                "fn": len(gold_keys - prediction_keys),
            }
        )
    return pd.DataFrame(rows)


def _has_valid_pubmed_identifier(row: pd.Series) -> bool:
    source = str(row.get("source", "")).strip().casefold()
    document_id = str(row.get("document_id", "")).strip()
    pmid = str(row.get("pmid", "")).strip()
    if source == "pubmed":
        return bool(re.fullmatch(r"\d{6,9}(?:\.0)?", pmid))
    if document_id.upper().startswith("PMID_"):
        return bool(re.fullmatch(r"\d{6,9}(?:\.0)?", pmid))
    return False


def _f1_from_array(array: np.ndarray) -> float:
    tp, fp, fn = array.sum(axis=0)
    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    return float(2 * precision * recall / (precision + recall) if precision + recall else 0.0)


def _replay_benchmark(
    data_dir: Path,
    output_dir: Path,
    *,
    seed: int,
    bootstrap_iterations: int,
) -> dict[str, Any]:
    sentences = pd.read_csv(data_dir / "benchmark_sentences.csv")
    gold_entities = pd.read_csv(data_dir / "normalized_gold_entities.csv")
    gold_relations = pd.read_csv(data_dir / "normalized_gold_relations.csv")
    entity_predictions = {
        "AISKG": pd.read_csv(data_dir / "aiskg_entities.csv"),
        "PubTator3": pd.read_csv(data_dir / "pubtator_entities.csv"),
        "StructuredLLM": pd.read_csv(data_dir / "llm_entities.csv"),
    }
    relation_predictions = {
        "AISKG": pd.read_csv(data_dir / "aiskg_relations.csv"),
        "StructuredLLM": pd.read_csv(data_dir / "llm_relations.csv"),
    }
    llm_validation = pd.read_csv(data_dir / "llm_validation_log.csv")
    source_manifest = _read_json(data_dir / "run_manifest.json")

    if len(sentences) != 150 or not sentences["benchmark_sentence_id"].is_unique:
        raise ValueError("Expected 150 unique benchmark sentences.")
    full_ids = sentences["benchmark_sentence_id"].astype(str).tolist()
    pubtator_eligible = sentences.apply(_has_valid_pubmed_identifier, axis=1)
    common_ids = sentences.loc[pubtator_eligible, "benchmark_sentence_id"].astype(str).tolist()
    ineligible_ids = sentences.loc[~pubtator_eligible, "benchmark_sentence_id"].astype(str).tolist()
    if len(common_ids) != 146 or len(full_ids) != 150:
        raise ValueError(f"Unexpected benchmark populations: common={len(common_ids)}, full={len(full_ids)}")
    if sorted(ineligible_ids) != sorted(source_manifest["pubtator_ineligible_sentence_ids"]):
        raise AssertionError("PubTator ineligible IDs do not match the corrected source manifest.")

    common_schema = sorted(set(gold_entities["type"]) & COMMON_ENTITY_TYPES)
    full_schema = sorted(set(gold_entities["type"]))
    count_store: dict[tuple[str, str, str, str], pd.DataFrame] = {}
    id_store: dict[tuple[str, str, str, str], list[str]] = {}
    metric_rows: list[dict[str, Any]] = []

    for mode in ("strict", "overlap"):
        for system, predictions in entity_predictions.items():
            counts = _entity_counts(gold_entities, predictions, common_schema, common_ids, mode)
            key = ("entity", "COMMON_146", mode, system)
            count_store[key] = counts
            id_store[key] = common_ids
            metric_rows.append(
                {
                    "task": "entity",
                    "schema": "COMMON_146",
                    "criterion": mode,
                    "system": system,
                    "n_sentences": len(common_ids),
                    **_prf_from_counts(counts),
                }
            )
        for system in ["AISKG", "StructuredLLM"]:
            predictions = entity_predictions[system]
            counts = _entity_counts(gold_entities, predictions, full_schema, full_ids, mode)
            key = ("entity", "FULL_DOMAIN_150", mode, system)
            count_store[key] = counts
            id_store[key] = full_ids
            metric_rows.append(
                {
                    "task": "entity",
                    "schema": "FULL_DOMAIN_150",
                    "criterion": mode,
                    "system": system,
                    "n_sentences": len(full_ids),
                    **_prf_from_counts(counts),
                }
            )

    for system, predictions in relation_predictions.items():
        counts = _relation_counts(gold_relations, predictions, full_ids)
        key = ("relation", "FULL_DOMAIN_150", "directed_strict", system)
        count_store[key] = counts
        id_store[key] = full_ids
        metric_rows.append(
            {
                "task": "relation",
                "schema": "FULL_DOMAIN_150",
                "criterion": "directed_strict",
                "system": system,
                "n_sentences": len(full_ids),
                **_prf_from_counts(counts),
            }
        )
    metrics = pd.DataFrame(metric_rows)

    rng = np.random.default_rng(seed)
    boot_values: dict[tuple[str, str, str, str], np.ndarray] = {}
    boot_indices: dict[tuple[str, ...], np.ndarray] = {}
    ci_rows: list[dict[str, Any]] = []
    for key, counts in count_store.items():
        ids = id_store[key]
        group = tuple(ids)
        if group not in boot_indices:
            boot_indices[group] = rng.integers(0, len(ids), size=(bootstrap_iterations, len(ids)))
        array = counts.set_index("benchmark_sentence_id").loc[ids, ["tp", "fp", "fn"]].to_numpy()
        values = np.asarray([_f1_from_array(array[index]) for index in boot_indices[group]], dtype=float)
        boot_values[key] = values
        ci_rows.append(
            {
                "task": key[0],
                "schema": key[1],
                "criterion": key[2],
                "system": key[3],
                "f1_ci_low": float(np.quantile(values, 0.025)),
                "f1_ci_high": float(np.quantile(values, 0.975)),
            }
        )
    metrics = metrics.merge(pd.DataFrame(ci_rows), on=["task", "schema", "criterion", "system"])

    comparisons = [
        ("entity", "COMMON_146", "strict", "AISKG", "PubTator3"),
        ("entity", "COMMON_146", "strict", "AISKG", "StructuredLLM"),
        ("entity", "COMMON_146", "strict", "StructuredLLM", "PubTator3"),
        ("relation", "FULL_DOMAIN_150", "directed_strict", "AISKG", "StructuredLLM"),
    ]
    paired_rows: list[dict[str, Any]] = []
    mcnemar_rows: list[dict[str, Any]] = []
    for task, schema, criterion, system_a, system_b in comparisons:
        key_a = (task, schema, criterion, system_a)
        key_b = (task, schema, criterion, system_b)
        ids = id_store[key_a]
        if ids != id_store[key_b]:
            raise AssertionError(f"Paired populations differ for {key_a} and {key_b}")
        differences = boot_values[key_a] - boot_values[key_b]
        paired_rows.append(
            {
                "task": task,
                "schema": schema,
                "criterion": criterion,
                "system_A": system_a,
                "system_B": system_b,
                "n_sentences": len(ids),
                "delta_f1_A_minus_B": float(np.median(differences)),
                "ci_low": float(np.quantile(differences, 0.025)),
                "ci_high": float(np.quantile(differences, 0.975)),
                "bootstrap_two_sided_p": float(
                    min(1.0, 2 * min((differences <= 0).mean(), (differences >= 0).mean()))
                ),
            }
        )
        counts_a = count_store[key_a].set_index("benchmark_sentence_id").loc[ids]
        counts_b = count_store[key_b].set_index("benchmark_sentence_id").loc[ids]
        exact_a = counts_a["fp"].eq(0) & counts_a["fn"].eq(0)
        exact_b = counts_b["fp"].eq(0) & counts_b["fn"].eq(0)
        a_correct_b_wrong = int((exact_a & ~exact_b).sum())
        a_wrong_b_correct = int((~exact_a & exact_b).sum())
        discordant = a_correct_b_wrong + a_wrong_b_correct
        exact_p = (
            float(binomtest(min(a_correct_b_wrong, a_wrong_b_correct), discordant, 0.5).pvalue)
            if discordant
            else 1.0
        )
        mcnemar_rows.append(
            {
                "task": task,
                "schema": schema,
                "criterion": criterion,
                "system_A": system_a,
                "system_B": system_b,
                "n_sentences": len(ids),
                "A_correct_B_wrong": a_correct_b_wrong,
                "A_wrong_B_correct": a_wrong_b_correct,
                "mcnemar_exact_p": exact_p,
            }
        )
    paired = pd.DataFrame(paired_rows)
    mcnemar = pd.DataFrame(mcnemar_rows)
    mcnemar["holm_p"] = multipletests(mcnemar["mcnemar_exact_p"], method="holm")[1]

    coverage_rows: list[dict[str, Any]] = []
    for entity_type in common_schema:
        for system, predictions in entity_predictions.items():
            counts = _entity_counts(gold_entities, predictions, [entity_type], common_ids, "strict")
            coverage_rows.append(
                {
                    "scope": "COMMON_146",
                    "entity_type": entity_type,
                    "system": system,
                    "n_sentences": len(common_ids),
                    "gold_n": int(
                        (
                            gold_entities["benchmark_sentence_id"].isin(common_ids)
                            & gold_entities["type"].eq(entity_type)
                        ).sum()
                    ),
                    "pred_n": int(
                        (
                            predictions["benchmark_sentence_id"].isin(common_ids)
                            & predictions["type"].eq(entity_type)
                        ).sum()
                    ),
                    **_prf_from_counts(counts),
                }
            )
    for entity_type in full_schema:
        for system in ["AISKG", "StructuredLLM"]:
            predictions = entity_predictions[system]
            counts = _entity_counts(gold_entities, predictions, [entity_type], full_ids, "strict")
            coverage_rows.append(
                {
                    "scope": "FULL_DOMAIN_150",
                    "entity_type": entity_type,
                    "system": system,
                    "n_sentences": len(full_ids),
                    "gold_n": int(gold_entities["type"].eq(entity_type).sum()),
                    "pred_n": int(predictions["type"].eq(entity_type).sum()),
                    **_prf_from_counts(counts),
                }
            )
    coverage = pd.DataFrame(coverage_rows)

    _assert_frame_close(
        metrics,
        pd.read_csv(data_dir / "system_metrics.csv"),
        ["task", "schema", "criterion", "system"],
        tolerance=1e-12,
    )
    _assert_frame_close(
        paired,
        pd.read_csv(data_dir / "paired_bootstrap_differences.csv"),
        ["task", "schema", "criterion", "system_A", "system_B"],
        tolerance=1e-12,
    )
    _assert_frame_close(
        mcnemar,
        pd.read_csv(data_dir / "mcnemar_holm_tests.csv"),
        ["task", "schema", "criterion", "system_A", "system_B"],
        tolerance=1e-12,
    )
    _assert_frame_close(
        coverage,
        pd.read_csv(data_dir / "entity_class_coverage.csv"),
        ["scope", "entity_type", "system"],
        tolerance=1e-12,
    )

    relation_row = metrics[
        metrics["task"].eq("relation")
        & metrics["system"].eq("AISKG")
        & metrics["criterion"].eq("directed_strict")
    ].iloc[0]
    if (int(relation_row.tp), int(relation_row.fp), int(relation_row.fn)) != (27, 0, 29):
        raise AssertionError("AISKG strict relation projection no longer matches the frozen locked rule.")
    if not llm_validation["json_valid"].astype(bool).all():
        raise AssertionError("Corrected structured-LLM run is not 100% JSON-valid.")
    if not paired["bootstrap_two_sided_p"].between(0.0, 1.0).all():
        raise AssertionError("A bootstrap p-value lies outside [0, 1].")
    pubtator_relation_table = pd.read_csv(data_dir / "pubtator_relations.csv")
    if len(pubtator_relation_table) != 0:
        raise AssertionError("The corrected source manifest states that PubTator supplied no usable relation objects.")
    if (
        (metrics["task"].eq("relation") & metrics["system"].eq("PubTator3")).any()
        or source_manifest.get("pubtator_relation_status") != "not_evaluable_no_relation_objects"
    ):
        raise AssertionError("PubTator relation performance must remain not evaluable, not zero.")

    output_dir.mkdir(parents=True, exist_ok=True)
    metrics.to_csv(output_dir / "system_metrics.csv", index=False)
    paired.to_csv(output_dir / "paired_bootstrap_differences.csv", index=False)
    mcnemar.to_csv(output_dir / "mcnemar_holm_tests.csv", index=False)
    coverage.to_csv(output_dir / "entity_class_coverage.csv", index=False)
    llm_validation.to_csv(output_dir / "llm_validation_log.csv", index=False)
    for filename in [
        "normalized_gold_entities.csv",
        "normalized_gold_relations.csv",
        "aiskg_entities.csv",
        "aiskg_relations.csv",
        "pubtator_entities.csv",
        "pubtator_relations.csv",
        "llm_entities.csv",
        "llm_relations.csv",
        "run_manifest.json",
    ]:
        shutil.copy2(data_dir / filename, output_dir / filename)

    common_plot = metrics[
        metrics["task"].eq("entity")
        & metrics["schema"].eq("COMMON_146")
        & metrics["criterion"].eq("strict")
    ]
    figure, axis = plt.subplots(figsize=(9, 5))
    x_positions = np.arange(len(common_plot))
    axis.bar(x_positions, common_plot["f1"].to_numpy())
    axis.errorbar(
        x_positions,
        common_plot["f1"].to_numpy(),
        yerr=np.vstack(
            [
                common_plot["f1"].to_numpy() - common_plot["f1_ci_low"].to_numpy(),
                common_plot["f1_ci_high"].to_numpy() - common_plot["f1"].to_numpy(),
            ]
        ),
        fmt="none",
        capsize=5,
    )
    axis.set_xticks(x_positions, common_plot["system"].tolist())
    axis.set_ylim(0, 1)
    axis.set_ylabel("Micro F1")
    axis.set_title("Common-schema strict entity extraction (n=146)")
    figure.tight_layout()
    figure.savefig(output_dir / "figure_common_schema_entity_f1.png", dpi=300, bbox_inches="tight")
    plt.close(figure)

    recall_pivot = coverage[coverage["scope"].eq("COMMON_146")].pivot(
        index="entity_type", columns="system", values="recall"
    )
    figure, axis = plt.subplots(figsize=(9, 6))
    image = axis.imshow(recall_pivot.to_numpy(), vmin=0, vmax=1, aspect="auto")
    axis.set_xticks(range(len(recall_pivot.columns)), recall_pivot.columns, rotation=20, ha="right")
    axis.set_yticks(range(len(recall_pivot.index)), recall_pivot.index)
    for row_index in range(len(recall_pivot.index)):
        for column_index in range(len(recall_pivot.columns)):
            axis.text(column_index, row_index, f"{recall_pivot.iloc[row_index, column_index]:.2f}", ha="center", va="center")
    axis.set_title("Strict entity recall by class (common subset, n=146)")
    figure.colorbar(image, ax=axis, label="Recall")
    figure.tight_layout()
    figure.savefig(output_dir / "figure_entity_class_recall.png", dpi=300, bbox_inches="tight")
    plt.close(figure)

    benchmark_workbook = output_dir / "AISKG_three_system_benchmark_reproduced_v3.1.2.xlsx"
    benchmark_sheets = {
        "Metrics": metrics,
        "Paired_bootstrap": paired,
        "McNemar_Holm": mcnemar,
        "Class_coverage": coverage,
        "LLM_validation": llm_validation,
    }
    benchmark_sheet_options = {
        "Metrics": {
            "width_overrides": {
                "task": 12,
                "schema": 20,
                "criterion": 18,
                "system": 20,
                "n_sentences": 14,
                "tp": 10,
                "fp": 10,
                "fn": 10,
                "precision": 14,
                "recall": 14,
                "f1": 14,
                "f1_ci_low": 14,
                "f1_ci_high": 14,
            },
            "max_width": 22,
        },
        "Paired_bootstrap": {"max_width": 30},
        "McNemar_Holm": {"max_width": 30},
        "Class_coverage": {"max_width": 25},
        "LLM_validation": {
            "width_overrides": {"sentence_id": 16, "json_valid": 14, "rejected_items": 16, "validation_note": 70},
            "wrap_columns": ["validation_note"],
            "max_width": 70,
        },
    }
    with pd.ExcelWriter(benchmark_workbook, engine="openpyxl") as writer:
        _set_fixed_workbook_properties(writer)
        for sheet_name, frame in benchmark_sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
        _style_excel_workbook(writer, benchmark_sheets, sheet_options=benchmark_sheet_options)
    _normalise_xlsx_archive(benchmark_workbook)

    manuscript_text = (data_dir / "manuscript_ready_results.txt").read_text(encoding="utf-8").strip()
    provenance_note = (
        manuscript_text
        + " The archived structured-LLM execution recorded model revision 'main' rather than an immutable "
        "weight commit; therefore the item-level results and statistics are exactly replayable, but a future "
        "live inference rerun is not claimed to be bit-for-bit identical."
    )
    (output_dir / "manuscript_ready_results_with_provenance_note.txt").write_text(provenance_note + "\n", encoding="utf-8")

    status = {
        "release": RELEASE_VERSION,
        "status": "CORRECTED_THREE_SYSTEM_STATISTICAL_REPLAY_PASSED",
        "population": {"common_schema_sentences": 146, "full_domain_sentences": 150},
        "pubtator_sentence_coverage": int(source_manifest["pubtator_sentence_coverage"]),
        "pubtator_relation_status": source_manifest["pubtator_relation_status"],
        "structured_llm_json_valid_n": int(llm_validation["json_valid"].astype(bool).sum()),
        "structured_llm_json_valid_rate": float(llm_validation["json_valid"].astype(bool).mean()),
        "llm_model_id": source_manifest["llm_model_id"],
        "llm_model_revision_requested": source_manifest["llm_model_revision"],
        "llm_model_revision_is_immutable_commit": bool(
            re.fullmatch(r"[0-9a-fA-F]{40,64}", str(source_manifest["llm_model_revision"]))
        ),
        "item_level_result_replay_exact": True,
        "future_live_llm_inference_bitwise_reproducible": False,
        "reference_tables_verified": [
            "system_metrics.csv",
            "paired_bootstrap_differences.csv",
            "mcnemar_holm_tests.csv",
            "entity_class_coverage.csv",
        ],
    }
    _write_json(status, output_dir / "BENCHMARK_REPRODUCIBILITY_STATUS.json")
    return {
        "metrics": metrics,
        "paired": paired,
        "mcnemar": mcnemar,
        "coverage": coverage,
        "status": status,
        "manuscript_text": provenance_note,
    }


def run_replay(
    data_root: str | Path,
    output_root: str | Path,
    *,
    seed: int = DEFAULT_SEED,
    pathway_bootstraps: int = DEFAULT_PATHWAY_BOOTSTRAPS,
    benchmark_bootstraps: int = DEFAULT_BENCHMARK_BOOTSTRAPS,
    clean: bool = True,
) -> ReplayResult:
    """Run the complete deterministic v3.1.2 additional-analysis replay.

    Parameters
    ----------
    data_root:
        Directory containing ``pathway/`` and ``benchmark/`` frozen inputs.
    output_root:
        Destination directory. A sibling deterministic ZIP is produced.
    seed:
        Locked analysis seed; values other than the v3.1.2 release seed are rejected.
    pathway_bootstraps:
        Locked number of overlap-aware unique-item pathway bootstrap replicates.
    benchmark_bootstraps:
        Locked number of sentence-cluster benchmark bootstrap replicates.
    clean:
        Must remain ``True`` for the frozen release; stale outputs are not permitted.
    """

    if seed != DEFAULT_SEED or pathway_bootstraps != DEFAULT_PATHWAY_BOOTSTRAPS or benchmark_bootstraps != DEFAULT_BENCHMARK_BOOTSTRAPS:
        raise ValueError(
            "The manuscript-facing v3.1.2 replay is locked to seed=20260817, "
            "pathway_bootstraps=10000, and benchmark_bootstraps=5000. "
            "Exploratory sensitivity analyses must be run separately and must not overwrite the frozen replay."
        )
    if not clean:
        raise ValueError("The manuscript-facing v3.1.2 replay requires clean=True to prevent stale-output contamination.")

    data_root = Path(data_root).resolve()
    output_root = Path(output_root).resolve()
    pathway_data = data_root / "pathway"
    benchmark_data = data_root / "benchmark"
    required_directories = [pathway_data, benchmark_data]
    missing_directories = [str(path) for path in required_directories if not path.is_dir()]
    if missing_directories:
        raise FileNotFoundError(f"Missing additional-analysis data directories: {missing_directories}")
    checksum_entries = _verify_input_checksums(data_root)
    if clean and output_root.exists():
        shutil.rmtree(output_root)
    output_root.mkdir(parents=True, exist_ok=True)

    pathway_output = output_root / "pathway_validation"
    benchmark_output = output_root / "benchmark"
    pathway_result = _replay_pathway(
        pathway_data,
        pathway_output,
        seed=seed,
        bootstrap_iterations=pathway_bootstraps,
    )
    benchmark_result = _replay_benchmark(
        benchmark_data,
        benchmark_output,
        seed=seed,
        bootstrap_iterations=benchmark_bootstraps,
    )

    source_upload_hashes_path = data_root / "source_upload_sha256.json"
    source_hashes = _read_json(source_upload_hashes_path) if source_upload_hashes_path.exists() else {}
    combined_manifest = {
        "release_version": RELEASE_VERSION,
        "base_frozen_framework_version": BASE_FROZEN_RELEASE,
        "audited_repository_commit": AUDITED_COMMIT,
        "analysis_seed": seed,
        "frozen_input_checksum_entries_verified": checksum_entries,
        "pathway_bootstrap_replicates": pathway_bootstraps,
        "benchmark_bootstrap_replicates": benchmark_bootstraps,
        "pathway": pathway_result["status"],
        "benchmark": benchmark_result["status"],
        "source_upload_sha256": source_hashes,
        "reporting_boundary": {
            "pathway_reviewer_level_replay": "fully reproducible from three public metadata-sanitized completed workbooks",
            "pathway_adjudication_replay": "all 92 required third-expert decisions independently validated",
            "pathway_final_endpoint_replay": "fully reproducible after reconstruction of all 805 final ratings",
            "corrected_benchmark_statistical_replay": "fully reproducible from archived item-level gold and predictions",
            "pubtator_relations": "not evaluable because the API returned no usable relation objects",
            "future_llm_live_rerun": "not claimed bitwise-identical because the executed run recorded revision main",
            "original_unsanitized_reviewer_files": "preserved only in the private provenance archive and excluded from public GitHub",
        },
    }
    manifest_path = output_root / "COMBINED_REPRODUCIBILITY_MANIFEST.json"
    _write_json(combined_manifest, manifest_path)

    reporting_status = f"""# AISKG v{RELEASE_VERSION} publication reporting status

## Reproduced and permitted for reporting

- Reviewer-level pathway validation: all 805 paired A/B ratings were reconstructed from the public completed workbooks; 22 direct disagreements and all 92 required third-expert adjudications were validated.
- Expanded pathway validation: 23/95 pre-refinement pathways and 26/52 outcome-aware refined pathways were completely correct.
- Absolute pathway-correctness difference: 25.8 percentage points; overlap-aware bootstrap 95% CI 14.4–37.5 percentage points.
- Corrected common-schema strict entity micro-F1 (146 PMID-eligible sentences): AISKG 0.904, PubTator3 0.501, StructuredLLM 0.503.
- Corrected full-domain directed strict relation micro-F1 (150 sentences): AISKG 0.651 and StructuredLLM 0.009.
- PubTator relation performance is not evaluable and must not be reported as zero.

## Provenance qualifications

- The public reviewer workbooks are content-equivalent, metadata-sanitized copies. The untouched source uploads are intentionally excluded from GitHub and retained only in the private provenance archive.
- Two Expert A complete-pathway cells (XPV-0052 and XPV-0074) differ from the deterministic component roll-up; both cases are preserved as submitted and resolved to `No` by the third-expert workbook.
- The archived LLM outputs are item-level reproducible and all 150 responses were parseable JSON after validation. The executed run recorded model revision `main`, so a future live rerun is not claimed to use identical model weights.
- The v3.0.0 frozen core pipeline remains the authoritative route for the original manuscript-snapshot outputs; v3.1.2 adds complete reviewer-level pathway replay and retains the corrected three-system benchmark replay introduced in v3.1.1.
"""
    (output_root / "PUBLICATION_REPORTING_STATUS.md").write_text(reporting_status, encoding="utf-8")

    success_marker = output_root / "ADDITIONAL_ANALYSES_SUCCESS.txt"
    success_marker.write_text(
        "AISKG v3.1.2 corrected additional analyses completed successfully.\n"
        "Pathway expected tables: PASS\n"
        "Corrected benchmark reference tables: PASS\n"
        "Invalid zero-score treatment for PubTator relations: excluded\n",
        encoding="utf-8",
    )

    checksum_path = output_root / "SHA256SUMS.txt"
    write_sha256s(output_root, checksum_path, exclude={checksum_path.name})
    output_archive = output_root.parent / "AISKG_v3.1.2_additional_analyses_reproduced.zip"
    deterministic_zip(output_root, output_archive)

    return ReplayResult(
        output_root=output_root,
        output_archive=output_archive,
        pathway_summary=pathway_output / "pathway_correctness_summary.csv",
        benchmark_metrics=benchmark_output / "system_metrics.csv",
        success_marker=success_marker,
        manifest=manifest_path,
    )
